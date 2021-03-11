#Import library
from openpyxl import *
from statistics import *
from textblob import TextBlob

#Load excel file
exFile=load_workbook("rest.xlsx")

#Confirm the active sheet
sheet1TL=exFile.active

#Make a list for star ratings from excel
starRatings=[]
for i in range (2,24):
    starRating=sheet1TL.cell(column=7,row=i).value
    starRatings.append(float(starRating))

#Make a list for review_text from excel
reviewList=[]
for i in range (2,24):
    reviewText=sheet1TL.cell(column=4,row=i).value
    reviewSentP=TextBlob(reviewText)
    reviewSentP=reviewSentP.sentiment.polarity
    reviewList.append(reviewSentP)

#Define a function
def calculation(a,b):
    print("Calulations for Stars:")
    print("Average Rating is ", round(sum(a)/len(a),2))
    print("Minimum Rating is ", min(a))
    print("Maximum Rating is ", max(a))
    print("Range of Rating is ", max(a)-min(a))
    print("Standard deviation of Rating is ", round(stdev(a),2),"\n")

    print("Calulations for Reviews:")
    print("Average is ", round(sum(b)/len(b),2))
    print("Minimum is ", round(min(b),2))
    print("Maximum is ", round(max(b),2))
    print("Range of is ", round((max(b)-min(b)),2))
    print("Standard deviation is ", round(stdev(b),2),"\n")

#Run the function
print("-----------------------Total-----------------------")
calculation(starRatings,reviewList)

#Make each list for Applebee's and Olive Garden 
aStars=[]
oStars=[]
aReviews=[]
oReviews=[]

#Read values from excel
for i in range (2,24):
    name=sheet1TL.cell(column=2,row=i).value
    if name=="Applebee's":
        aStar=sheet1TL.cell(column=7,row=i).value
        aStars.append(float(aStar))
        aReview=sheet1TL.cell(column=4,row=i).value
        aReviewSentP=TextBlob(aReview)
        aReviewSentP=aReviewSentP.sentiment.polarity
        aReviews.append(aReviewSentP)

    elif name=="Olive Garden":
        oStar=sheet1TL.cell(column=7,row=i).value
        oStars.append(float(oStar))
        oReview=sheet1TL.cell(column=4,row=i).value
        oReviewSentP=TextBlob(oReview)
        oReviewSentP=oReviewSentP.sentiment.polarity
        oReviews.append(oReviewSentP)

#Print calculation for each
print("--------------------Applebee's--------------------")
calculation(aStars,aReviews)

print("-------------------Olive Garden-------------------")
calculation(oStars,oReviews)

#Summary
print("""
Average rating of Olive Garden is higher than average rating of Applebee's.
Applebee's doesn't have 5star rating while Olive Garden has it.
Considering that the range of polarity is from -1 to 1(negative to positive),
Olive Garden has more positive reviews than Applebee's because Olive Garden's
average of polarity is greater than Applebee's's
""")

